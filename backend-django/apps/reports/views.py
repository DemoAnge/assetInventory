"""
Módulo de Reportería — KPIs, estadísticas y exportación CSV/Excel.

Reportes disponibles:
  - Inventario general          /export/inventory/         (CSV + Excel)
  - SEPS activos fijos          /export/seps/              (CSV + Excel)
  - Historial de bajas          /export/bajas/             (CSV + Excel)
  - Historial de movimientos    /export/movements/         (CSV + Excel)
  - Tabla de depreciación       /export/depreciation/      (CSV + Excel)
"""
import csv
import io
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.assets.models import Asset
from apps.movements.models import AssetMovement, MovementType
from apps.shared.permissions import IsAnyStaff, IsAdmin

# ── Constantes compartidas ────────────────────────────────────────────────────

SEPS_DESC = {
    "1801": "Terrenos",
    "1802": "Edificios y locales",
    "1803": "Maquinaria y equipo",
    "1804": "Muebles y enseres",
    "1805": "Equipos de cómputo",
    "1806": "Vehículos",
    "1807": "Equipos de comunicación",
    "1899": "Otros activos fijos",
}

CATEGORY_TO_SEPS = {
    "COMPUTO":          "1805",
    "VEHICULO":         "1806",
    "MAQUINARIA":       "1803",
    "MUEBLE":           "1804",
    "INMUEBLE":         "1801",
    "TELECOMUNICACION": "1807",
    "OTRO":             "1899",
}

MOVEMENT_TYPE_LABELS = {
    "TRASLADO":     "Traslado",
    "PRESTAMO":     "Préstamo",
    "DEVOLUCION":   "Devolución",
    "REASIGNACION": "Reasignación",
    "INGRESO":      "Ingreso",
    "BAJA":         "Baja / Retiro",
    "REACTIVACION": "Reactivación",
}

_INVENTORY_HEADERS = [
    "Código", "Nombre", "Categoría", "Estado", "Marca", "Modelo", "Serie",
    "Agencia", "Departamento", "Área", "Custodio",
    "Fecha compra", "Proveedor", "Factura",
    "Vida útil (años)", "Tasa dep. (%)", "Cuenta SEPS",
    "F. garantía", "En mantenimiento", "TI crítico",
    "Activo principal", "Tipo componente", "QR UUID", "Creado en",
]

_SEPS_HEADERS = [
    "Cuenta SEPS", "Descripción cuenta",
    "Código activo", "Nombre", "N° serie", "Categoría", "Estado",
    "Agencia", "Custodio", "Proveedor", "N° Factura",
    "Fecha compra", "Fecha activación", "Vida útil (años)", "Tasa dep. (%)",
    "Valor de compra ($)", "Dep. acumulada ($)", "Valor en libros ($)", "Valor residual ($)",
    "F. garantía", "Tot. depreciado", "Req. mantenimiento",
]

_BAJAS_HEADERS = [
    "Fecha baja", "Código activo", "Nombre", "N° serie", "Categoría",
    "Agencia origen", "Depto. origen", "Área origen", "Custodio origen",
    "Motivo", "Autorizado por", "N° documento", "Registrado por", "Fecha registro",
]

_MOVEMENTS_HEADERS = [
    "Fecha", "Tipo de movimiento",
    "Código activo", "Nombre", "N° serie", "Categoría",
    "Agencia origen", "Depto. origen", "Área origen", "Custodio origen",
    "Agencia destino", "Depto. destino", "Área destino", "Custodio destino",
    "Motivo", "Autorizado por", "N° documento", "Arrastre", "Registrado por", "Fecha registro",
]

_DEPRECIATION_HEADERS = [
    "Cuenta SEPS", "Descripción cuenta",
    "Código activo", "Nombre", "N° serie", "Categoría", "Estado", "Agencia",
    "Fecha compra", "Años desde compra", "Vida útil (años)", "Tasa dep. (%)",
    "Valor de compra ($)", "Dep. anual ($)", "Dep. acumulada ($)",
    "Valor en libros ($)", "Valor residual ($)",
    "Tot. depreciado",
]


# ── Utilidades ────────────────────────────────────────────────────────────────

def _dec(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _fmt(value) -> str:
    try:
        return str(_dec(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return "0.00"


def _seps_code(asset) -> str:
    return asset.seps_account_code or CATEGORY_TO_SEPS.get(asset.category, "1899")


def _years_elapsed(purchase_date) -> float:
    if not purchase_date:
        return 0.0
    delta = date.today() - purchase_date
    return round(delta.days / 365.25, 2)


def _parse_date(request, key):
    """Parsea fecha del querystring en formato YYYY-MM-DD."""
    raw = request.query_params.get(key, "")
    if not raw:
        return None
    try:
        from datetime import datetime
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


# ── Filas de datos ────────────────────────────────────────────────────────────

def _asset_row(asset) -> list:
    brand     = asset.asset_model.brand.name if asset.asset_model else ""
    model_name = asset.asset_model.name if asset.asset_model else ""
    custodian = str(asset.custodian) if asset.custodian else ""
    return [
        asset.asset_code, asset.name,
        asset.get_category_display(), asset.get_status_display(),
        brand, model_name, asset.serial_number or "",
        asset.agency.name if asset.agency else "",
        asset.department.name if asset.department else "",
        asset.area.name if asset.area else "",
        custodian,
        str(asset.purchase_date or ""), asset.supplier or "", asset.invoice_number or "",
        asset.useful_life_years or "", str(asset.depreciation_rate or ""),
        asset.seps_account_code or "",
        str(asset.warranty_expiry or ""),
        "Sí" if asset.requires_maintenance else "No",
        "Sí" if asset.is_critical_it else "No",
        asset.parent_asset.asset_code if asset.parent_asset else "",
        asset.get_component_type_display() if asset.component_type else "",
        str(asset.qr_uuid), str(asset.created_at.date()),
    ]


def _seps_row(asset) -> list:
    code = _seps_code(asset)
    return [
        code, SEPS_DESC.get(code, "Otros activos fijos"),
        asset.asset_code, asset.name,
        asset.serial_number or "",
        asset.get_category_display(), asset.get_status_display(),
        asset.agency.name if asset.agency else "",
        str(asset.custodian) if asset.custodian else "",
        asset.supplier or "", asset.invoice_number or "",
        str(asset.purchase_date or ""), str(asset.activation_date or ""),
        asset.useful_life_years or "", _fmt(asset.depreciation_rate),
        _fmt(asset.purchase_value), _fmt(asset.accumulated_depreciation),
        _fmt(asset.current_value), _fmt(asset.residual_value),
        str(asset.warranty_expiry or ""),
        "Sí" if asset.is_fully_depreciated else "No",
        "Sí" if asset.requires_maintenance else "No",
    ]


def _baja_row(mov) -> list:
    asset = mov.asset
    return [
        str(mov.movement_date),
        asset.asset_code if asset else "",
        asset.name if asset else "",
        asset.serial_number or "" if asset else "",
        asset.get_category_display() if asset else "",
        mov.origin_agency.name if mov.origin_agency else "",
        mov.origin_department.name if mov.origin_department else "",
        mov.origin_area.name if mov.origin_area else "",
        str(mov.origin_custodian) if mov.origin_custodian else "",
        mov.reason or "",
        str(mov.authorized_by) if mov.authorized_by else "",
        mov.document_ref or "",
        str(mov.created_by) if mov.created_by else "",
        str(mov.created_at.date()),
    ]


def _movement_row(mov) -> list:
    asset = mov.asset
    return [
        str(mov.movement_date),
        MOVEMENT_TYPE_LABELS.get(mov.movement_type, mov.movement_type),
        asset.asset_code if asset else "",
        asset.name if asset else "",
        asset.serial_number or "" if asset else "",
        asset.get_category_display() if asset else "",
        mov.origin_agency.name if mov.origin_agency else "",
        mov.origin_department.name if mov.origin_department else "",
        mov.origin_area.name if mov.origin_area else "",
        str(mov.origin_custodian) if mov.origin_custodian else "",
        mov.dest_agency.name if mov.dest_agency else "",
        mov.dest_department.name if mov.dest_department else "",
        mov.dest_area.name if mov.dest_area else "",
        str(mov.dest_custodian) if mov.dest_custodian else "",
        mov.reason or "",
        str(mov.authorized_by) if mov.authorized_by else "",
        mov.document_ref or "",
        "Sí" if mov.is_cascade else "No",
        str(mov.created_by) if mov.created_by else "",
        str(mov.created_at.date()),
    ]


def _depreciation_row(asset) -> list:
    code       = _seps_code(asset)
    years_gone = _years_elapsed(asset.purchase_date)
    dep_annual = _dec(asset.purchase_value) * _dec(asset.depreciation_rate) / Decimal("100")
    return [
        code, SEPS_DESC.get(code, "Otros activos fijos"),
        asset.asset_code, asset.name,
        asset.serial_number or "",
        asset.get_category_display(), asset.get_status_display(),
        asset.agency.name if asset.agency else "",
        str(asset.purchase_date or ""),
        years_gone,
        asset.useful_life_years or "",
        _fmt(asset.depreciation_rate),
        _fmt(asset.purchase_value),
        _fmt(dep_annual),
        _fmt(asset.accumulated_depreciation),
        _fmt(asset.current_value),
        _fmt(asset.residual_value),
        "Sí" if asset.is_fully_depreciated else "No",
    ]


# ── Querysets ─────────────────────────────────────────────────────────────────

def _inventory_queryset(request):
    qs = Asset.objects.select_related(
        "asset_model__brand", "asset_model__asset_type",
        "agency", "department", "area", "custodian", "parent_asset",
    ).order_by("asset_code")
    if c := request.query_params.get("category"): qs = qs.filter(category=c)
    if s := request.query_params.get("status"):   qs = qs.filter(status=s)
    if a := request.query_params.get("agency"):   qs = qs.filter(agency_id=a)
    return qs


def _seps_queryset(request):
    qs = Asset.objects.select_related(
        "agency", "department", "area", "custodian", "parent_asset", "asset_model__brand",
    ).filter(parent_asset__isnull=True).order_by("seps_account_code", "category", "asset_code")
    if a := request.query_params.get("agency"):   qs = qs.filter(agency_id=a)
    if c := request.query_params.get("category"): qs = qs.filter(category=c)
    if acc := request.query_params.get("account"):
        cats = [cat for cat, s in CATEGORY_TO_SEPS.items() if s == acc]
        qs = qs.filter(Q(seps_account_code=acc) | Q(category__in=cats))
    return qs


def _bajas_queryset(request):
    qs = AssetMovement.objects.select_related(
        "asset", "asset__agency",
        "origin_agency", "origin_department", "origin_area", "origin_custodian",
        "authorized_by", "created_by",
    ).filter(movement_type=MovementType.BAJA).order_by("-movement_date", "-created_at")

    date_from = _parse_date(request, "date_from")
    date_to   = _parse_date(request, "date_to")
    if date_from: qs = qs.filter(movement_date__gte=date_from)
    if date_to:   qs = qs.filter(movement_date__lte=date_to)
    if a := request.query_params.get("agency"):   qs = qs.filter(origin_agency_id=a)
    if c := request.query_params.get("category"): qs = qs.filter(asset__category=c)
    return qs


def _movements_queryset(request):
    qs = AssetMovement.objects.select_related(
        "asset",
        "origin_agency", "origin_department", "origin_area", "origin_custodian",
        "dest_agency", "dest_department", "dest_area", "dest_custodian",
        "authorized_by", "created_by",
    ).order_by("-movement_date", "-created_at")

    date_from = _parse_date(request, "date_from")
    date_to   = _parse_date(request, "date_to")
    if date_from: qs = qs.filter(movement_date__gte=date_from)
    if date_to:   qs = qs.filter(movement_date__lte=date_to)
    if mt := request.query_params.get("movement_type"): qs = qs.filter(movement_type=mt)
    if a  := request.query_params.get("agency"):
        qs = qs.filter(Q(origin_agency_id=a) | Q(dest_agency_id=a))
    if request.query_params.get("exclude_cascade") == "true":
        qs = qs.filter(is_cascade=False)
    return qs


def _depreciation_queryset(request):
    qs = Asset.objects.select_related(
        "agency", "asset_model__brand",
    ).filter(parent_asset__isnull=True).order_by("seps_account_code", "category", "asset_code")
    if c := request.query_params.get("category"): qs = qs.filter(category=c)
    if a := request.query_params.get("agency"):   qs = qs.filter(agency_id=a)
    if y := request.query_params.get("year"):
        try:
            qs = qs.filter(purchase_date__year=int(y))
        except ValueError:
            pass
    return qs


# ════════════════════════════════════════════════════════════════════════════════
# HELPERS EXCEL
# ════════════════════════════════════════════════════════════════════════════════

def _xl_styles():
    """Devuelve un dict con los estilos openpyxl más usados."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="CCCCCC")
    return {
        "hdr_fill":     PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid"),
        "hdr_font":     Font(name="Calibri", color="FFFFFF", bold=True, size=10),
        "hdr_align":    Alignment(horizontal="center", vertical="center", wrap_text=True),
        "total_fill":   PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"),
        "total_font":   Font(name="Calibri", color="FFFFFF", bold=True, size=10),
        "sub_fill":     PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
        "sub_font":     Font(name="Calibri", bold=True, size=10, color="1E3A8A"),
        "grp_fill":     PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"),
        "grp_font":     Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "zebra_fill":   PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
        "data_font":    Font(name="Calibri", size=10),
        "cell_border":  Border(left=thin, right=thin, top=thin, bottom=thin),
        "center":       Alignment(horizontal="center", vertical="center"),
        "left":         Alignment(vertical="center"),
        "right":        Alignment(horizontal="right", vertical="center"),
        "money_fmt":    '#,##0.00',
    }


def _write_header_row(ws, headers, st, row=1):
    from openpyxl.styles import Alignment
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = st["hdr_fill"]; cell.font = st["hdr_font"]
        cell.alignment = st["hdr_align"]; cell.border = st["cell_border"]
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = f"A{row + 1}"


def _autowidth(ws, max_col=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, max_col)


def _xl_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    r = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    r["Content-Disposition"] = f'attachment; filename="{filename}"'
    return r


# ════════════════════════════════════════════════════════════════════════════════
# VISTAS
# ════════════════════════════════════════════════════════════════════════════════

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        user      = request.user
        assets_qs = Asset.objects.filter(is_active=True)

        total_assets     = assets_qs.count()
        total_components = assets_qs.filter(parent_asset__isnull=False).count()
        critical_it      = assets_qs.filter(is_critical_it=True).count()
        fully_dep        = assets_qs.filter(is_fully_depreciated=True).count()
        needs_maint      = assets_qs.filter(requires_maintenance=True).count()

        by_category = list(assets_qs.values("category").annotate(count=Count("id")).order_by("-count"))
        by_status   = list(assets_qs.values("status").annotate(count=Count("id")).order_by("-count"))
        by_agency   = list(
            assets_qs.filter(agency__isnull=False)
            .values("agency__name").annotate(count=Count("id")).order_by("-count")[:10]
        )

        from apps.audit.models import AuditLog, AuditAction
        recent_ingresos = list(
            AuditLog.objects.filter(action=AuditAction.ASSET_ACTIVATION)
            .order_by("-action_date")[:5]
            .values("object_code", "object_name", "action_date", "user_email")
        )

        financial = {}
        if user.role in ("ADMIN", "CONTABILIDAD"):
            financial = {
                "total_assets_with_value": assets_qs.count(),
                "fully_depreciated_count": fully_dep,
                "sales_count": 0,
            }

        return Response({
            "total_assets":      total_assets,
            "total_components":  total_components,
            "critical_it":       critical_it,
            "fully_deprecated":  fully_dep,
            "needs_maintenance": needs_maint,
            "alerts_unresolved": 0,
            "by_category":       by_category,
            "by_status":         by_status,
            "by_agency":         by_agency,
            "recent_activity":   recent_ingresos,
            "financial":         financial,
            "generated_at":      timezone.now().isoformat(),
        })


class AssetsByMonthView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        from django.db.models.functions import TruncMonth
        from datetime import timedelta
        cutoff = timezone.now().date() - timedelta(days=365)
        data = (
            Asset.objects.filter(created_at__date__gte=cutoff)
            .annotate(month=TruncMonth("created_at"))
            .values("month").annotate(count=Count("id")).order_by("month")
        )
        return Response([{"month": str(r["month"])[:7], "count": r["count"]} for r in data])


# ── Inventario ────────────────────────────────────────────────────────────────

class ExportInventoryCSVView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        qs = _inventory_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="inventario_activos.csv"'
        response.write("﻿")
        w = csv.writer(response)
        w.writerow(_INVENTORY_HEADERS)
        for asset in qs:
            w.writerow(_asset_row(asset))
        return response


class ExportInventoryExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        import openpyxl
        qs = _inventory_queryset(request)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Inventario Activos"
        st = _xl_styles()
        _write_header_row(ws, _INVENTORY_HEADERS, st)
        for row_idx, asset in enumerate(qs, 2):
            for col_idx, val in enumerate(_asset_row(asset), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = st["data_font"]; cell.border = st["cell_border"]
                cell.alignment = st["left"]
                if row_idx % 2 == 0: cell.fill = st["zebra_fill"]
        _autowidth(ws)
        return _xl_response(wb, "inventario_activos.xlsx")


# ── SEPS ──────────────────────────────────────────────────────────────────────

class ExportSEPSCSVView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        qs = _seps_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="reporte_seps_activos_fijos.csv"'
        response.write("﻿")
        w = csv.writer(response)
        w.writerow(_SEPS_HEADERS)

        current_code = None
        gt = {k: Decimal(0) for k in ("pv", "ad", "cv", "rv")}
        gg = {k: Decimal(0) for k in gt}

        def _flush(code, totals):
            blank = [""] * len(_SEPS_HEADERS)
            blank[1] = f"SUBTOTAL {code} — {SEPS_DESC.get(code, '')}"
            blank[15] = _fmt(totals["pv"]); blank[16] = _fmt(totals["ad"])
            blank[17] = _fmt(totals["cv"]); blank[18] = _fmt(totals["rv"])
            w.writerow(blank); w.writerow([])

        for asset in qs:
            code = _seps_code(asset)
            if current_code and code != current_code:
                _flush(current_code, gt)
                gt = {k: Decimal(0) for k in gt}
            current_code = code
            w.writerow(_seps_row(asset))
            for key, field in [("pv", "purchase_value"), ("ad", "accumulated_depreciation"),
                                ("cv", "current_value"), ("rv", "residual_value")]:
                v = _dec(getattr(asset, field))
                gt[key] += v; gg[key] += v

        if current_code:
            _flush(current_code, gt)

        blank = [""] * len(_SEPS_HEADERS)
        blank[1] = "TOTAL GENERAL"
        blank[15] = _fmt(gg["pv"]); blank[16] = _fmt(gg["ad"])
        blank[17] = _fmt(gg["cv"]); blank[18] = _fmt(gg["rv"])
        w.writerow(blank)
        return response


class ExportSEPSExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        import openpyxl
        qs = list(_seps_queryset(request))
        wb = openpyxl.Workbook()
        st = _xl_styles()

        # ── Hoja resumen ──
        ws_sum = wb.active; ws_sum.title = "Resumen SEPS"
        from openpyxl.styles import Font
        ws_sum["A1"] = "REPORTE SEPS — ACTIVOS FIJOS"
        ws_sum["A1"].font = Font(name="Calibri", bold=True, size=13, color="1E3A8A")
        ws_sum["A2"] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws_sum["A2"].font = Font(name="Calibri", size=9, color="6B7280")

        sum_hdrs = ["Cuenta SEPS", "Descripción", "N° activos",
                    "Valor de compra ($)", "Dep. acumulada ($)", "Valor en libros ($)", "Valor residual ($)"]
        _write_header_row(ws_sum, sum_hdrs, st, row=4)

        grouped = {}
        for asset in qs:
            code = _seps_code(asset)
            if code not in grouped:
                grouped[code] = {"n": 0, "pv": Decimal(0), "ad": Decimal(0), "cv": Decimal(0), "rv": Decimal(0)}
            g = grouped[code]
            g["n"] += 1
            g["pv"] += _dec(asset.purchase_value); g["ad"] += _dec(asset.accumulated_depreciation)
            g["cv"] += _dec(asset.current_value);  g["rv"] += _dec(asset.residual_value)

        gg = {"n": 0, "pv": Decimal(0), "ad": Decimal(0), "cv": Decimal(0), "rv": Decimal(0)}
        sr = 5
        for code in sorted(grouped):
            g = grouped[code]
            for col_idx, val in enumerate(
                [code, SEPS_DESC.get(code, "Otros"), g["n"],
                 float(g["pv"]), float(g["ad"]), float(g["cv"]), float(g["rv"])], 1
            ):
                cell = ws_sum.cell(row=sr, column=col_idx, value=val)
                cell.font = st["data_font"]; cell.border = st["cell_border"]
                cell.alignment = st["right"] if col_idx > 2 else st["left"]
                if col_idx >= 4: cell.number_format = st["money_fmt"]
            for k in gg: gg[k] += g[k]
            sr += 1

        for col_idx, val in enumerate(
            ["", "TOTAL GENERAL", gg["n"],
             float(gg["pv"]), float(gg["ad"]), float(gg["cv"]), float(gg["rv"])], 1
        ):
            cell = ws_sum.cell(row=sr, column=col_idx, value=val)
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
            cell.alignment = st["right"] if col_idx > 2 else st["left"]
            if col_idx >= 4: cell.number_format = st["money_fmt"]
        _autowidth(ws_sum)

        # ── Hoja detalle ──
        ws = wb.create_sheet("Detalle por cuenta")
        _write_header_row(ws, _SEPS_HEADERS, st)
        FIN = {16, 17, 18, 19}
        N = len(_SEPS_HEADERS)

        sorted_assets = sorted(qs, key=lambda a: (_seps_code(a), a.asset_code))
        row_idx = 2; current_code = None; group_buf = []

        def _flush_group(code, buf):
            nonlocal row_idx
            for col in range(1, N + 1):
                cell = ws.cell(row=row_idx, column=col,
                               value=f"  Cuenta {code} — {SEPS_DESC.get(code, '')}" if col == 1 else "")
                cell.fill = st["grp_fill"]; cell.font = st["grp_font"]
                cell.border = st["cell_border"]; cell.alignment = st["left"]
            ws.row_dimensions[row_idx].height = 18; row_idx += 1

            gt = {k: Decimal(0) for k in ("pv", "ad", "cv", "rv")}
            for nth, a in enumerate(buf):
                for ci, val in enumerate(_seps_row(a), 1):
                    cell = ws.cell(row=row_idx, column=ci, value=val)
                    cell.font = st["data_font"]; cell.border = st["cell_border"]
                    cell.alignment = st["right"] if ci in FIN else st["left"]
                    if nth % 2 == 0: cell.fill = st["zebra_fill"]
                    if ci in FIN: cell.number_format = st["money_fmt"]
                gt["pv"] += _dec(a.purchase_value); gt["ad"] += _dec(a.accumulated_depreciation)
                gt["cv"] += _dec(a.current_value);  gt["rv"] += _dec(a.residual_value)
                row_idx += 1

            for ci in range(1, N + 1):
                v = ""
                if ci == 2: v = f"Subtotal {code} ({len(buf)} activos)"
                elif ci == 16: v = float(gt["pv"])
                elif ci == 17: v = float(gt["ad"])
                elif ci == 18: v = float(gt["cv"])
                elif ci == 19: v = float(gt["rv"])
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.fill = st["sub_fill"]; cell.font = st["sub_font"]
                cell.border = st["cell_border"]
                cell.alignment = st["right"] if ci in FIN else st["left"]
                if ci in FIN: cell.number_format = st["money_fmt"]
            ws.row_dimensions[row_idx].height = 18; row_idx += 2

        for asset in sorted_assets:
            code = _seps_code(asset)
            if current_code and code != current_code:
                _flush_group(current_code, group_buf); group_buf = []
            current_code = code; group_buf.append(asset)
        if group_buf: _flush_group(current_code, group_buf)

        # Total general
        for ci in range(1, N + 1):
            v = ""
            if ci == 2: v = f"TOTAL GENERAL ({gg['n']} activos)"
            elif ci == 16: v = float(gg["pv"])
            elif ci == 17: v = float(gg["ad"])
            elif ci == 18: v = float(gg["cv"])
            elif ci == 19: v = float(gg["rv"])
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
            cell.alignment = st["right"] if ci in FIN else st["left"]
            if ci in FIN: cell.number_format = st["money_fmt"]
        ws.row_dimensions[row_idx].height = 22
        _autowidth(ws)
        wb.active = ws_sum
        return _xl_response(wb, "reporte_seps_activos_fijos.xlsx")


# ── Historial de bajas ────────────────────────────────────────────────────────

class ExportBajasCSVView(APIView):
    """GET /api/v1/reports/export/bajas/ — CSV historial de bajas."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        qs = _bajas_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="historial_bajas.csv"'
        response.write("﻿")
        w = csv.writer(response)
        w.writerow(_BAJAS_HEADERS)
        for mov in qs:
            w.writerow(_baja_row(mov))
        return response


class ExportBajasExcelView(APIView):
    """GET /api/v1/reports/export/bajas/excel/ — Excel historial de bajas."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font

        qs = list(_bajas_queryset(request))
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Historial de Bajas"
        st = _xl_styles()

        # Título
        ws["A1"] = "HISTORIAL DE BAJAS — ACTIVOS FIJOS"
        ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="7F1D1D")
        ws["A2"] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Total registros: {len(qs)}"
        ws["A2"].font = Font(name="Calibri", size=9, color="6B7280")
        ws.merge_cells("A2:N2")

        _write_header_row(ws, _BAJAS_HEADERS, st, row=3)
        ws.freeze_panes = "A4"

        # Colores por categoría
        cat_colors = {
            "COMPUTO":          "DBEAFE", "VEHICULO":     "DCFCE7",
            "MAQUINARIA":       "FEF9C3", "MUEBLE":       "FEE2E2",
            "TELECOMUNICACION": "F3E8FF", "INMUEBLE":     "FFEDD5",
            "OTRO":             "F1F5F9",
        }
        from openpyxl.styles import PatternFill

        for row_idx, mov in enumerate(qs, 4):
            row_data = _baja_row(mov)
            cat = mov.asset.category if mov.asset else ""
            fill_color = cat_colors.get(cat, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = st["data_font"]; cell.border = st["cell_border"]
                cell.alignment = st["left"]; cell.fill = fill

        # Fila total
        total_row = len(qs) + 4
        for ci in range(1, len(_BAJAS_HEADERS) + 1):
            cell = ws.cell(row=total_row, column=ci,
                           value=f"Total: {len(qs)} bajas registradas" if ci == 1 else "")
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
        ws.row_dimensions[total_row].height = 20

        _autowidth(ws)
        return _xl_response(wb, "historial_bajas.xlsx")


# ── Historial de movimientos ──────────────────────────────────────────────────

class ExportMovementsCSVView(APIView):
    """GET /api/v1/reports/export/movements/ — CSV historial de movimientos."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        qs = _movements_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="historial_movimientos.csv"'
        response.write("﻿")
        w = csv.writer(response)
        w.writerow(_MOVEMENTS_HEADERS)
        for mov in qs:
            w.writerow(_movement_row(mov))
        return response


class ExportMovementsExcelView(APIView):
    """GET /api/v1/reports/export/movements/excel/ — Excel historial de movimientos agrupado."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        qs = list(_movements_queryset(request))
        wb = openpyxl.Workbook()
        st = _xl_styles()

        # ── Hoja resumen por tipo ──
        ws_sum = wb.active; ws_sum.title = "Resumen"
        ws_sum["A1"] = "HISTORIAL DE MOVIMIENTOS — RESUMEN"
        ws_sum["A1"].font = Font(name="Calibri", bold=True, size=13, color="1E3A8A")
        ws_sum["A2"] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(qs)} movimientos"
        ws_sum["A2"].font = Font(name="Calibri", size=9, color="6B7280")

        _write_header_row(ws_sum, ["Tipo de movimiento", "N° movimientos"], st, row=4)
        type_counts = {}
        for mov in qs:
            label = MOVEMENT_TYPE_LABELS.get(mov.movement_type, mov.movement_type)
            type_counts[label] = type_counts.get(label, 0) + 1

        sr = 5
        for label, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            ws_sum.cell(row=sr, column=1, value=label).font = st["data_font"]
            ws_sum.cell(row=sr, column=2, value=count).font = st["data_font"]
            for ci in range(1, 3):
                ws_sum.cell(row=sr, column=ci).border = st["cell_border"]
                ws_sum.cell(row=sr, column=ci).alignment = st["left"] if ci == 1 else st["right"]
            sr += 1
        for ci in range(1, 3):
            cell = ws_sum.cell(row=sr, column=ci,
                               value="TOTAL" if ci == 1 else len(qs))
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
            cell.alignment = st["left"] if ci == 1 else st["right"]
        _autowidth(ws_sum)

        # ── Hoja detalle completo ──
        ws = wb.create_sheet("Detalle")
        ws["A1"] = "HISTORIAL COMPLETO DE MOVIMIENTOS"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="1E3A8A")
        _write_header_row(ws, _MOVEMENTS_HEADERS, st, row=2)
        ws.freeze_panes = "A3"

        # Colores por tipo de movimiento
        type_colors = {
            "TRASLADO":     "DBEAFE", "PRESTAMO":     "FEF9C3",
            "DEVOLUCION":   "DCFCE7", "REASIGNACION": "F3E8FF",
            "INGRESO":      "D1FAE5", "BAJA":         "FEE2E2",
            "REACTIVACION": "ECFDF5",
        }

        for row_idx, mov in enumerate(qs, 3):
            fill_color = type_colors.get(mov.movement_type, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_idx, val in enumerate(_movement_row(mov), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = st["data_font"]; cell.border = st["cell_border"]
                cell.alignment = st["left"]; cell.fill = fill

        # ── Hojas por tipo ──
        for mt_value, mt_label in MOVEMENT_TYPE_LABELS.items():
            subset = [m for m in qs if m.movement_type == mt_value]
            if not subset:
                continue
            ws_t = wb.create_sheet(mt_label[:25])
            ws_t["A1"] = f"{mt_label.upper()} — {len(subset)} registros"
            ws_t["A1"].font = Font(name="Calibri", bold=True, size=11, color="1E3A8A")
            _write_header_row(ws_t, _MOVEMENTS_HEADERS, st, row=2)
            ws_t.freeze_panes = "A3"
            fill = PatternFill(
                start_color=type_colors.get(mt_value, "FFFFFF"),
                end_color=type_colors.get(mt_value, "FFFFFF"),
                fill_type="solid",
            )
            for row_idx, mov in enumerate(subset, 3):
                for col_idx, val in enumerate(_movement_row(mov), 1):
                    cell = ws_t.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = st["data_font"]; cell.border = st["cell_border"]
                    cell.alignment = st["left"]; cell.fill = fill
            _autowidth(ws_t)

        _autowidth(ws)
        wb.active = ws_sum
        return _xl_response(wb, "historial_movimientos.xlsx")


# ── Tabla de depreciación ─────────────────────────────────────────────────────

class ExportDepreciationCSVView(APIView):
    """GET /api/v1/reports/export/depreciation/ — CSV tabla de depreciación LORTI."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        qs = _depreciation_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="tabla_depreciacion.csv"'
        response.write("﻿")
        w = csv.writer(response)
        w.writerow(_DEPRECIATION_HEADERS)

        gt = {k: Decimal(0) for k in ("pv", "dep_a", "ad", "cv", "rv")}
        current_code = None
        group_totals = {k: Decimal(0) for k in gt}

        def _flush_dep(code, totals):
            blank = [""] * len(_DEPRECIATION_HEADERS)
            blank[1] = f"SUBTOTAL {code} — {SEPS_DESC.get(code, '')}"
            blank[12] = _fmt(totals["pv"]);   blank[13] = _fmt(totals["dep_a"])
            blank[14] = _fmt(totals["ad"]);   blank[15] = _fmt(totals["cv"])
            blank[16] = _fmt(totals["rv"])
            w.writerow(blank); w.writerow([])

        for asset in qs:
            code = _seps_code(asset)
            if current_code and code != current_code:
                _flush_dep(current_code, group_totals)
                group_totals = {k: Decimal(0) for k in gt}
            current_code = code
            w.writerow(_depreciation_row(asset))
            dep_a = _dec(asset.purchase_value) * _dec(asset.depreciation_rate) / Decimal("100")
            group_totals["pv"]    += _dec(asset.purchase_value)
            group_totals["dep_a"] += dep_a
            group_totals["ad"]    += _dec(asset.accumulated_depreciation)
            group_totals["cv"]    += _dec(asset.current_value)
            group_totals["rv"]    += _dec(asset.residual_value)
            gt["pv"]    += _dec(asset.purchase_value);  gt["dep_a"] += dep_a
            gt["ad"]    += _dec(asset.accumulated_depreciation)
            gt["cv"]    += _dec(asset.current_value);   gt["rv"]    += _dec(asset.residual_value)

        if current_code:
            _flush_dep(current_code, group_totals)

        blank = [""] * len(_DEPRECIATION_HEADERS)
        blank[1] = "TOTAL GENERAL"
        blank[12] = _fmt(gt["pv"]);   blank[13] = _fmt(gt["dep_a"])
        blank[14] = _fmt(gt["ad"]);   blank[15] = _fmt(gt["cv"])
        blank[16] = _fmt(gt["rv"])
        w.writerow(blank)
        return response


class ExportDepreciationExcelView(APIView):
    """GET /api/v1/reports/export/depreciation/excel/ — Excel tabla de depreciación."""
    permission_classes = [IsAuthenticated, IsAnyStaff]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        qs = list(_depreciation_queryset(request))
        wb = openpyxl.Workbook()
        st = _xl_styles()

        # ── Hoja resumen ──
        ws_sum = wb.active; ws_sum.title = "Resumen Depreciación"
        ws_sum["A1"] = "TABLA DE DEPRECIACIÓN — LORTI ART. 28"
        ws_sum["A1"].font = Font(name="Calibri", bold=True, size=13, color="065F46")
        ws_sum["A2"] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws_sum["A2"].font = Font(name="Calibri", size=9, color="6B7280")

        sum_hdrs = ["Cuenta SEPS", "Descripción", "N° activos", "Tasa promedio (%)",
                    "Valor de compra ($)", "Dep. anual ($)", "Dep. acumulada ($)",
                    "Valor en libros ($)", "Valor residual ($)", "Tot. depreciados"]
        _write_header_row(ws_sum, sum_hdrs, st, row=4)

        grouped_dep = {}
        for asset in qs:
            code = _seps_code(asset)
            if code not in grouped_dep:
                grouped_dep[code] = {
                    "n": 0, "rates": [], "pv": Decimal(0), "dep_a": Decimal(0),
                    "ad": Decimal(0), "cv": Decimal(0), "rv": Decimal(0), "fully": 0,
                }
            g = grouped_dep[code]
            g["n"] += 1
            if asset.depreciation_rate: g["rates"].append(float(asset.depreciation_rate))
            dep_a = _dec(asset.purchase_value) * _dec(asset.depreciation_rate) / Decimal("100")
            g["pv"] += _dec(asset.purchase_value);  g["dep_a"] += dep_a
            g["ad"] += _dec(asset.accumulated_depreciation)
            g["cv"] += _dec(asset.current_value);   g["rv"] += _dec(asset.residual_value)
            if asset.is_fully_depreciated: g["fully"] += 1

        gg = {"n": 0, "pv": Decimal(0), "dep_a": Decimal(0),
              "ad": Decimal(0), "cv": Decimal(0), "rv": Decimal(0), "fully": 0}
        sr = 5
        for code in sorted(grouped_dep):
            g = grouped_dep[code]
            avg_rate = round(sum(g["rates"]) / len(g["rates"]), 2) if g["rates"] else 0
            for ci, val in enumerate(
                [code, SEPS_DESC.get(code, "Otros"), g["n"], avg_rate,
                 float(g["pv"]), float(g["dep_a"]), float(g["ad"]),
                 float(g["cv"]), float(g["rv"]), g["fully"]], 1
            ):
                cell = ws_sum.cell(row=sr, column=ci, value=val)
                cell.font = st["data_font"]; cell.border = st["cell_border"]
                cell.alignment = st["right"] if ci > 2 else st["left"]
                if ci in {5, 6, 7, 8, 9}: cell.number_format = st["money_fmt"]
            for k in gg:
                if k != "n" and k != "fully": gg[k] += g[k]
            gg["n"] += g["n"]; gg["fully"] += g["fully"]
            sr += 1

        for ci, val in enumerate(
            ["", "TOTAL GENERAL", gg["n"], "",
             float(gg["pv"]), float(gg["dep_a"]), float(gg["ad"]),
             float(gg["cv"]), float(gg["rv"]), gg["fully"]], 1
        ):
            cell = ws_sum.cell(row=sr, column=ci, value=val)
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
            cell.alignment = st["right"] if ci > 2 else st["left"]
            if ci in {5, 6, 7, 8, 9}: cell.number_format = st["money_fmt"]
        _autowidth(ws_sum)

        # ── Hoja detalle ──
        ws = wb.create_sheet("Detalle Depreciación")
        ws["A1"] = "DETALLE DE DEPRECIACIÓN POR ACTIVO — LORTI ART. 28"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="065F46")
        _write_header_row(ws, _DEPRECIATION_HEADERS, st, row=2)
        ws.freeze_panes = "A3"

        FIN = {13, 14, 15, 16, 17}
        N   = len(_DEPRECIATION_HEADERS)
        row_idx = 3; current_code = None; group_buf = []

        def _flush_dep_xl(code, buf):
            nonlocal row_idx
            for ci in range(1, N + 1):
                cell = ws.cell(row=row_idx, column=ci,
                               value=f"  Cuenta {code} — {SEPS_DESC.get(code, '')}" if ci == 1 else "")
                cell.fill = st["grp_fill"]; cell.font = st["grp_font"]
                cell.border = st["cell_border"]; cell.alignment = st["left"]
            ws.row_dimensions[row_idx].height = 18; row_idx += 1

            gt2 = {k: Decimal(0) for k in ("pv", "dep_a", "ad", "cv", "rv")}
            for nth, a in enumerate(buf):
                dep_a = _dec(a.purchase_value) * _dec(a.depreciation_rate) / Decimal("100")
                fully_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") \
                    if a.is_fully_depreciated else (
                        st["zebra_fill"] if nth % 2 == 0 else None
                    )
                for ci, val in enumerate(_depreciation_row(a), 1):
                    cell = ws.cell(row=row_idx, column=ci, value=val)
                    cell.font = st["data_font"]; cell.border = st["cell_border"]
                    cell.alignment = st["right"] if ci in FIN else st["left"]
                    if fully_fill: cell.fill = fully_fill
                    if ci in FIN: cell.number_format = st["money_fmt"]
                gt2["pv"] += _dec(a.purchase_value); gt2["dep_a"] += dep_a
                gt2["ad"] += _dec(a.accumulated_depreciation)
                gt2["cv"] += _dec(a.current_value); gt2["rv"] += _dec(a.residual_value)
                row_idx += 1

            for ci in range(1, N + 1):
                v = ""
                if ci == 2: v = f"Subtotal {code} ({len(buf)} activos)"
                elif ci == 13: v = float(gt2["pv"])
                elif ci == 14: v = float(gt2["dep_a"])
                elif ci == 15: v = float(gt2["ad"])
                elif ci == 16: v = float(gt2["cv"])
                elif ci == 17: v = float(gt2["rv"])
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.fill = st["sub_fill"]; cell.font = st["sub_font"]
                cell.border = st["cell_border"]
                cell.alignment = st["right"] if ci in FIN else st["left"]
                if ci in FIN: cell.number_format = st["money_fmt"]
            ws.row_dimensions[row_idx].height = 18; row_idx += 2

        sorted_assets = sorted(qs, key=lambda a: (_seps_code(a), a.asset_code))
        for asset in sorted_assets:
            code = _seps_code(asset)
            if current_code and code != current_code:
                _flush_dep_xl(current_code, group_buf); group_buf = []
            current_code = code; group_buf.append(asset)
        if group_buf: _flush_dep_xl(current_code, group_buf)

        for ci in range(1, N + 1):
            v = ""
            if ci == 2: v = f"TOTAL GENERAL ({gg['n']} activos)"
            elif ci == 13: v = float(gg["pv"])
            elif ci == 14: v = float(gg["dep_a"])
            elif ci == 15: v = float(gg["ad"])
            elif ci == 16: v = float(gg["cv"])
            elif ci == 17: v = float(gg["rv"])
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.fill = st["total_fill"]; cell.font = st["total_font"]
            cell.border = st["cell_border"]
            cell.alignment = st["right"] if ci in FIN else st["left"]
            if ci in FIN: cell.number_format = st["money_fmt"]
        ws.row_dimensions[row_idx].height = 22
        _autowidth(ws)

        # Leyenda en detalle
        leg_row = row_idx + 2
        ws.cell(row=leg_row, column=1,
                value="* Filas en rojo = activos totalmente depreciados (valor en libros = valor residual)").font = \
            Font(name="Calibri", size=8, color="6B7280", italic=True)

        wb.active = ws_sum
        return _xl_response(wb, "tabla_depreciacion.xlsx")


# Alias de compatibilidad
ExportSEPSView = ExportSEPSCSVView
